import openpyxl as xl
from openpyxl.chart import BarChart,Reference 


wb= xl.load_workbook('transactions.xlsx')
sheet1=wb["Sheet1"]

cell1=sheet1["a1"]
sheet1.cell(1,1)

# print(cell1.value)
# print(sheet1.max_row)

for row in range(2,sheet1.max_row+1):
    cell=sheet1.cell(row,3)
    corrected_price=cell.value*0.9
    corrected_price_cell=sheet1.cell(row,4)
    corrected_price_cell.value=corrected_price

values = Reference(sheet1,
                min_col=4,
                min_row=2,
                max_col=4,
                max_row=sheet1.max_row)

product_id_values=Reference(sheet1,
                min_col=2,
                min_row=2,
                max_col=2,
                max_row=sheet1.max_row)

corrected_price_chart=BarChart()
corrected_price_chart.add_data(values)
sheet1.add_chart(corrected_price_chart,"e2")

product_id_chart=BarChart()
product_id_chart.add_data(product_id_values)
sheet1.add_chart(product_id_chart,"g2")


wb.save("transactions2.xlsx")